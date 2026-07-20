# -*- coding: utf-8 -*-
import base64
import csv
import io
import logging
from collections import OrderedDict
from datetime import datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools.mail import email_normalize

_logger = logging.getLogger(__name__)

# As colunas do template. A chave é o campo; o valor é o cabeçalho que a pessoa
# vê na planilha. Aceita-se qualquer ordem e maiúsculas/minúsculas -- ninguém
# vai manter a ordem de uma planilha que passou por três pessoas.
COLUMNS = [
    # Primeira coluna e opcional: uma planilha sem ela importa para a lista
    # escolhida na tela, como sempre foi. Com ela, um arquivo só resolve
    # dezenas de listas -- que é o caso real de quem exporta um histórico
    # inteiro (as SOBs do Odoo 15 saem com 132 títulos de uma vez).
    ('list', "Lista"),
    ('name', "Nome"),
    ('email', "E-mail"),
    ('phone', "Telefone"),
    ('bonus_partner_type', "Tipo"),
    ('street', "Endereço"),
    ('city', "Cidade"),
    ('zip', "CEP"),
    ('note', "Observação"),
    # Coluna de LISTA, não de contato, viajando numa planilha de contatos: a
    # data se repete em todas as linhas do mesmo título e vale a maior. Fica
    # aqui porque o arquivo é chato e único -- criar uma segunda aba só para
    # uma data por lista custaria mais a quem preenche do que a repetição.
    ('last_shipment', "Último envio"),
    # Quem a pessoa É ("Crítica literária da Folha") -- vai para o campo do
    # CONTATO, não para a nota do vínculo: a bio vale em toda lista, a
    # Observação é desta lista.
    ('bio', "Quem é"),
]
# Sinônimos que aparecem em planilha de verdade (a que vem do assessor, a que
# alguém exportou do Mailchimp). Mapear cinco variações aqui poupa a pessoa de
# renomear coluna à mão, que é onde a importação costuma morrer.
ALIASES = {
    'list': ('lista', 'list', 'título', 'titulo', 'livro', 'obra',
             'lista vip'),
    'name': ('nome', 'name', 'contato', 'nome completo'),
    'email': ('e-mail', 'email', 'e mail', 'correio', 'mail'),
    'phone': ('telefone', 'phone', 'celular', 'fone', 'whatsapp', 'tel'),
    'bonus_partner_type': ('tipo', 'type', 'categoria', 'perfil'),
    'street': ('endereço', 'endereco', 'street', 'rua', 'logradouro'),
    'city': ('cidade', 'city', 'município', 'municipio'),
    'zip': ('cep', 'zip', 'código postal', 'codigo postal'),
    'note': ('observação', 'observacao', 'note', 'obs', 'notas', 'veículo',
             'veiculo', 'meio'),
    # SEM 'data'/'date' genéricos: planilha de verdade (Mailchimp, assessor)
    # tem coluna "Data" que é data de cadastro -- viraria "último envio" da
    # lista em silêncio, e a gravação só-para-frente tornava o estrago
    # permanente. Sinônimo de coluna só quando o nome diz DO QUE é a data.
    'last_shipment': ('último envio', 'ultimo envio', 'last shipment',
                      'data do envio', 'data do último envio',
                      'data do ultimo envio', 'última remessa',
                      'ultima remessa'),
    # SEM 'perfil': já é sinônimo de Tipo ali em cima, e a mesma palavra em
    # dois campos faz a coluna alimentar o errado -- calada.
    'bio': ('quem é', 'quem e', 'bio', 'sobre', 'descrição',
            'descricao', 'who they are'),
}
TYPE_WORDS = {
    'author': ('autor', 'autora', 'author', 'escritor', 'escritora'),
    'journalist': ('jornalista', 'journalist', 'imprensa', 'crítico', 'critico'),
    'influencer': ('influencer', 'influenciador', 'influenciadora', 'blogueiro',
                   'blogueira', 'booktuber'),
    'bookshop': ('livraria', 'bookshop', 'livreiro', 'loja'),
    'other': ('outro', 'outra', 'other'),
}


class ProductBonusListImport(models.TransientModel):
    """Subir uma planilha e virar lista VIP.

    O TODO chamava isto de "o buraco mais feio", e o buraco não é ler o
    arquivo: é NÃO criar duzentos contatos duplicados. Uma lista de imprensa
    quase sempre é gente que já está na base -- e um import ingênuo transforma
    "Ana Prado <ana@folha.com>" num segundo cadastro, que quebra o histórico de
    bonificações, o Score e a checagem de quem já recebeu o título.

    Daí o desenho: PRÉVIA antes de gravar, casamento por e-mail normalizado (e
    por nome só se a pessoa mandar), e relatório do que aconteceu linha a linha.
    """
    _name = 'product.bonus.list.import'
    _description = 'Import Contacts into a VIP List'

    file = fields.Binary(string="Planilha", required=True, attachment=False)
    filename = fields.Char()

    # Um campo só, com "Criar «Nome»" no próprio dropdown -- o padrão do Odoo
    # para escolher-ou-criar. Dois campos ("Lista existente" + "Nova lista")
    # obrigavam a pessoa a entender a mecânica antes de fazer a coisa.
    # Deixou de ser obrigatório quando a planilha traz coluna "Lista": exigir
    # aqui uma lista que o arquivo já nomeia obrigaria a pessoa a escolher uma
    # das 132 só para o formulário deixar passar.
    list_id = fields.Many2one(
        'product.bonus.list', string="Lista",
        domain=[('active', '=', True)],
        help="Só para planilha de uma lista. Se o arquivo tiver coluna "
             "'Lista', ela manda -- e este campo é ignorado.")

    match_by_name = fields.Boolean(
        string="Casar por nome quando não houver e-mail", default=True,
        help="Sem e-mail, tenta achar um contato com o mesmo nome exato. "
             "Desligue se a base tem muitos homônimos -- na dúvida o "
             "importador cria um contato novo, que é o erro mais fácil de "
             "desfazer.")
    create_missing = fields.Boolean(
        string="Criar quem não existe", default=True,
        help="Desligado, a importação só liga à lista quem já está na base e "
             "relata os demais -- útil para conferir uma planilha antes de "
             "deixá-la entrar na base de contatos.")

    # O modelo fica anexado aqui, pronto para baixar -- não atrás de um botão
    # que gera na hora. Quem abre a tela vê o exemplo antes de precisar
    # perguntar como é o formato.
    template_file = fields.Binary(
        string="Planilha de exemplo", readonly=True,
        default=lambda self: base64.b64encode(self._build_template()))
    template_filename = fields.Char(
        readonly=True, default="modelo-lista-vip.xlsx")

    state = fields.Selection(
        [('upload', 'upload'), ('preview', 'preview'), ('done', 'done')],
        default='upload')
    preview_html = fields.Html(readonly=True, sanitize=False)
    result_html = fields.Html(readonly=True, sanitize=False)

    # ------------------------------------------------------------------
    # leitura
    # ------------------------------------------------------------------
    def _read_rows(self):
        """A planilha em dicionários {campo: valor}, cabeçalho já traduzido."""
        self.ensure_one()
        raw = base64.b64decode(self.file or b'')
        if not raw:
            raise UserError(_("Nenhum arquivo."))
        name = (self.filename or '').lower()
        if name.endswith(('.xlsx', '.xlsm')):
            rows = self._read_xlsx(raw)
        else:
            rows = self._read_csv(raw)
        if not rows:
            raise UserError(_("A planilha está vazia."))

        header, body = rows[0], rows[1:]
        mapping = {}
        for index, cell in enumerate(header):
            key = (cell or '').strip().lower()
            for field, names in ALIASES.items():
                # Primeiro-ganha: com "Tipo" e outra coluna sinônima no mesmo
                # arquivo, a da direita sobrescrevia o mapeamento e os dados da
                # esquerda sumiam sem aviso.
                if key in names and field not in mapping:
                    mapping[field] = index
                    break
        if 'name' not in mapping and 'email' not in mapping:
            raise UserError(_(
                "Não achei as colunas. A planilha precisa de pelo menos "
                "'Nome' ou 'E-mail' no cabeçalho.\n\n"
                "Baixe o modelo no botão ao lado e use os mesmos títulos."))
        if 'list' not in mapping and not self.list_id:
            raise UserError(_(
                "Para onde vão estes contatos? Escolha uma lista no campo "
                "acima, ou ponha uma coluna 'Lista' na planilha dizendo, "
                "linha a linha, a que lista cada contato pertence."))

        out = []
        for line in body:
            record = {}
            for field, index in mapping.items():
                value = line[index] if index < len(line) else None
                record[field] = str(value).strip() if value not in (None, '') else ''
            if any(record.values()):
                out.append(record)
        return out

    def _read_csv(self, raw):
        try:
            text = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            # Planilha salva no Excel brasileiro costuma sair em latin-1; morrer
            # por acento seria o motivo mais bobo de a importação falhar.
            text = raw.decode('latin-1')
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,\t')
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ';' if sample.count(';') > sample.count(',') else ','
        return [row for row in csv.reader(io.StringIO(text), dialect)]

    def _read_xlsx(self, raw):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_(
                "Este servidor não lê .xlsx. Salve a planilha como CSV."))
        book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        # Pela aba chamada "Contatos" quando ela existe: quem preenche o modelo
        # costuma deixar a aba de instruções aberta ao salvar, e aí a "ativa"
        # seria a errada.
        sheet = next(
            (ws for ws in book.worksheets if (ws.title or '').strip().lower()
             in ('contatos', 'contacts')),
            book.active)
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]

    # ------------------------------------------------------------------
    # casamento
    # ------------------------------------------------------------------
    def _list_name(self, row):
        """A que lista esta linha pertence. Coluna manda; campo da tela supre."""
        return (row.get('list') or '').strip() or (self.list_id.name or '')

    @staticmethod
    def _person_key(row):
        return (email_normalize(row.get('email') or '')
                or (row.get('name') or '').strip().lower())

    def _list_targets(self, decisions):
        """{nome da lista: lista existente ou vazio}, na ordem da planilha.

        Casar pelo nome, inclusive arquivada: recriar uma lista que a pessoa
        arquivou de propósito seria desfazer uma decisão dela em silêncio.
        """
        BList = self.env['product.bonus.list'].with_context(active_test=False)
        names = []
        for decision in decisions:
            if decision['kind'] in ('match', 'create') \
                    and decision['list_name'] not in names:
                names.append(decision['list_name'])
        found = {b.name: b for b in BList.search([('name', 'in', names)])}
        return OrderedDict((n, found.get(n, BList.browse())) for n in names)

    def _resolve(self, rows):
        """Decide, para cada linha, quem ela é: existente, nova ou repetida.

        Devolve a lista de decisões -- nada é gravado aqui, porque a prévia usa
        exatamente este resultado. Prever e executar por caminhos diferentes é
        como se produz a prévia que mente.

        A repetição é medida DENTRO de cada lista, não no arquivo inteiro: num
        arquivo multi-lista a mesma pessoa aparecer em vinte títulos é o normal
        (é o histórico dela), e um dedup global a jogaria em um só, calado.
        """
        Partner = self.env['res.partner']
        decisions = []
        seen = set()          # (lista, pessoa) -- repetido de verdade
        resolved = {}         # pessoa -> parceiro, para não buscar vinte vezes

        for row in rows:
            email = email_normalize(row.get('email') or '') or ''
            name = (row.get('name') or '').strip()
            key = email or name.lower()
            target = self._list_name(row)
            decision = {'row': row, 'list_name': target, 'partner': None}

            if not target:
                decisions.append(dict(decision, kind='no_list'))
                continue
            if key and (target, key) in seen:
                decisions.append(dict(decision, kind='dup_file'))
                continue
            if key:
                seen.add((target, key))

            if key in resolved:
                partner = resolved[key]
            else:
                partner = Partner.browse()
                if email:
                    partner = Partner.search(
                        [('email_normalized', '=', email)], limit=1)
                if not partner and name and self.match_by_name and not email:
                    partner = Partner.search([('name', '=ilike', name)], limit=1)
                if key:
                    resolved[key] = partner

            if partner:
                decisions.append(dict(decision, kind='match', partner=partner))
            elif not name:
                decisions.append(dict(decision, kind='no_name'))
            elif self.create_missing:
                decisions.append(dict(decision, kind='create'))
            else:
                decisions.append(dict(decision, kind='skipped'))
        return decisions

    @staticmethod
    def _parse_date(value):
        """A data como ela chega de planilha de verdade, ou False.

        O leitor transforma toda célula em texto, então uma data do xlsx chega
        como "2026-05-12 00:00:00" e uma digitada à mão como "12/05/2026". Uma
        data ilegível NÃO é erro de importação: a lista entra sem data, porque
        recusar 800 contatos por causa de uma célula mal digitada seria a troca
        errada.
        """
        text = (value or '').strip()
        if not text:
            return False
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y',
                    '%d-%m-%Y'):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return False

    def _list_dates(self, decisions):
        """{nome da lista: maior data encontrada}. Vale a mais recente porque a
        coluna é por contato e a pergunta é da lista: "quando este título saiu
        pela última vez"."""
        dates = {}
        for decision in decisions:
            if decision['kind'] not in ('match', 'create'):
                continue
            day = self._parse_date(decision['row'].get('last_shipment'))
            if not day:
                continue
            current = dates.get(decision['list_name'])
            if current is None or day > current:
                dates[decision['list_name']] = day
        return dates

    @staticmethod
    def _partner_type(word):
        key = (word or '').strip().lower()
        for value, words in TYPE_WORDS.items():
            if key in words:
                return value
        return False

    # ------------------------------------------------------------------
    # ações
    # ------------------------------------------------------------------
    def action_preview(self):
        self.ensure_one()
        decisions = self._resolve(self._read_rows())
        counts = {}
        for decision in decisions:
            counts[decision['kind']] = counts.get(decision['kind'], 0) + 1

        # Sem endereço o pacote não sai do depósito -- melhor saber agora, com a
        # planilha na mão, do que no dia da expedição.
        incomplete = sum(
            1 for d in decisions
            if d['kind'] in ('create', 'match')
            and not (d['row'].get('street') and d['row'].get('city')
                     and d['row'].get('zip')))

        # Uma pessoa em vinte listas vira vinte linhas 'create', mas UM contato
        # novo. Contar linhas aqui prometeria vinte cadastros e assustaria com
        # razão.
        novos = {self._person_key(d['row']) for d in decisions if d['kind'] == 'create'}
        alvos = self._list_targets(decisions)

        lines = [
            ("%s vínculos com contatos que já estão na base" % counts.get('match', 0), 'ok'),
            ("%s contatos novos" % len(novos), 'ok'),
        ]
        if len(alvos) > 1 or any(d['row'].get('list') for d in decisions):
            criar = [n for n, blist in alvos.items() if not blist]
            lines.insert(0, (
                "%s listas: %s já existem, %s serão criadas"
                % (len(alvos), len(alvos) - len(criar), len(criar)), 'ok'))
            if criar:
                amostra = ", ".join(sorted(criar)[:6])
                lines.append(("Listas novas: %s%s"
                              % (amostra, "..." if len(criar) > 6 else ""), 'warn'))
        # A prévia mostra TUDO que o import vai gravar -- inclusive as datas de
        # lista, que antes só apareciam depois do fato.
        dates = self._list_dates(decisions)
        if dates:
            lines.append((
                "%s lista(s) receberão data de último envio (a maior da "
                "planilha; nunca volta para trás)" % len(dates), 'ok'))
        if counts.get('no_list'):
            lines.append(("%s linhas sem lista (ignoradas)" % counts['no_list'], 'warn'))
        if counts.get('skipped'):
            lines.append(("%s fora da base e não serão criados" % counts['skipped'], 'warn'))
        if counts.get('dup_file'):
            lines.append(("%s repetidos dentro da planilha (ignorados)" % counts['dup_file'], 'warn'))
        if counts.get('no_name'):
            lines.append(("%s sem nome nem e-mail (ignorados)" % counts['no_name'], 'warn'))
        if incomplete:
            lines.append(("%s sem endereço completo — o pacote não sai do "
                          "depósito sem rua, cidade e CEP" % incomplete, 'warn'))

        html = "<ul>%s</ul>" % "".join(
            '<li class="%s">%s</li>'
            % ('text-muted' if kind == 'warn' else '', text)
            for text, kind in lines)
        self.write({'preview_html': html, 'state': 'preview'})
        return self._reopen()

    def action_import(self):
        self.ensure_one()
        Partner = self.env['res.partner']
        Member = self.env['product.bonus.list.member']
        BList = self.env['product.bonus.list']
        decisions = self._resolve(self._read_rows())

        targets = self._list_targets(decisions)
        dates = self._list_dates(decisions)
        novas = []
        for name, blist in targets.items():
            if not blist:
                targets[name] = BList.create({
                    'name': name,
                    'last_shipment_on': dates.get(name) or False,
                })
                novas.append(name)
            elif dates.get(name) and (not blist.last_shipment_on
                                      or dates[name] > blist.last_shipment_on):
                # Só para frente: reimportar um arquivo antigo por cima não
                # pode fazer a lista parecer mais fria do que ela é.
                blist.last_shipment_on = dates[name]

        # active_test=False: quem saiu da lista continua existindo, e recriar
        # bateria na restrição de unicidade.
        existing = {}
        for member in Member.with_context(active_test=False).search(
                [('list_id', 'in', [b.id for b in targets.values()])]):
            existing[(member.list_id.id, member.partner_id.id)] = member

        # A mesma pessoa em vinte listas é UM cadastro. Sem este cache, a
        # segunda linha dela não acha o parceiro (a busca correu antes de ele
        # existir) e cria um duplicado -- justamente o que este importador
        # existe para evitar.
        criados = {}
        created = linked = revived = already = 0
        por_lista = {}
        for decision in decisions:
            if decision['kind'] not in ('match', 'create'):
                continue
            row = decision['row']
            blist = targets[decision['list_name']]
            key = self._person_key(row)
            partner = decision['partner'] or criados.get(key)
            if not partner:
                partner = Partner.create({
                    'name': row['name'],
                    'email': row.get('email') or False,
                    'phone': row.get('phone') or False,
                    'street': row.get('street') or False,
                    'city': row.get('city') or False,
                    'zip': row.get('zip') or False,
                    'is_bonus_recipient': True,
                    'bonus_partner_type': self._partner_type(
                        row.get('bonus_partner_type')),
                    'bonus_bio': row.get('bio') or False,
                })
                created += 1
                if key:
                    criados[key] = partner
            else:
                # Só preenche o VAZIO: o que foi digitado no Odoo é de quem
                # conhece a pessoa; a planilha não sobrescreve calada.
                updates = {}
                if row.get('bio') and not partner.bonus_bio:
                    updates['bonus_bio'] = row['bio']
                tipo = self._partner_type(row.get('bonus_partner_type'))
                if tipo and not partner.bonus_partner_type:
                    updates['bonus_partner_type'] = tipo
                if updates:
                    partner.write(updates)

            member = existing.get((blist.id, partner.id))
            if member is None:
                Member.create({
                    'list_id': blist.id,
                    'partner_id': partner.id,
                    'note': row.get('note') or False,
                })
                existing[(blist.id, partner.id)] = True
                linked += 1
                por_lista[blist.id] = por_lista.get(blist.id, 0) + 1
            elif member is not True and not member.active:
                member.write({'active': True, 'left_on': False})
                revived += 1
            else:
                already += 1

        for blist in targets.values():
            blist.message_post(body=_(
                "Importação de %(file)s: %(linked)s contatos ligados a esta "
                "lista (%(created)s cadastros novos no arquivo inteiro).",
                file=self.filename or _("planilha"),
                linked=por_lista.get(blist.id, 0), created=created))

        resumo = ("<p><b>%s</b> contatos criados · <b>%s</b> vínculos · "
                  "<b>%s</b> reativados · <b>%s</b> já estavam.</p>"
                  % (created, linked, revived, already))
        if len(targets) > 1:
            resumo += ("<p><b>%s</b> listas alimentadas, sendo <b>%s</b> "
                       "criadas agora.</p>" % (len(targets), len(novas)))
        self.write({'state': 'done', 'result_html': resumo})

        # Uma lista só: abre nela, como sempre. Várias: abre a tela de listas
        # filtrada, porque escolher uma das 132 para "a" tela seria arbitrário.
        if len(targets) == 1:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'product.bonus.list',
                'res_id': list(targets.values())[0].id,
                'view_mode': 'form',
            }
        return {
            'type': 'ir.actions.act_window',
            'name': _("Listas importadas"),
            'res_model': 'product.bonus.list',
            'view_mode': 'list,form',
            'domain': [('id', 'in', [b.id for b in targets.values()])],
        }

    @api.model
    def _build_template(self):
        """A planilha de exemplo, gerada pelo MESMO código que a lê.

        Modelo escrito à mão num arquivo estático envelhece calado: alguém muda
        um cabeçalho aqui e o exemplo passa a ensinar o formato errado. Gerando
        a partir de COLUMNS, os dois nunca divergem -- e o teste do ida-e-volta
        (gera, sobe, importa) garante que continue assim.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            _logger.warning("openpyxl ausente: sem planilha de exemplo")
            return b''

        book = Workbook()
        sheet = book.active
        sheet.title = "Contatos"
        sheet.append([label for _field, label in COLUMNS])
        head = PatternFill("solid", fgColor="DDDDDD")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = head
            cell.alignment = Alignment(vertical="center")
        # Ana em duas listas de propósito: é o exemplo que ensina o formato
        # multi-lista sem precisar de parágrafo explicando.
        for row in (
            ["Imprensa literária", "Ana Prado", "ana.prado@exemplo.com.br",
             "+55 11 90000-0000", "jornalista", "R. Augusta, 100", "São Paulo",
             "01305-000", "Caderno de cultura", "03/04/2026",
             "Crítica literária, caderno de cultura"],
            ["Imprensa literária", "Jorge Elias", "jorge@exemplo.com.br",
             "+55 21 90000-0000", "influencer", "Av. Atlântica, 500",
             "Rio de Janeiro", "22010-000", "Booktuber, 80 mil seguidores",
             "03/04/2026", "Booktuber, 80 mil seguidores"],
            ["Lançamento Transposição", "Ana Prado",
             "ana.prado@exemplo.com.br", "+55 11 90000-0000", "jornalista",
             "R. Augusta, 100", "São Paulo", "01305-000", "Caderno de cultura",
             "12/05/2026", "Crítica literária, caderno de cultura"],
            ["Lançamento Transposição", "Marta Reis", "marta@exemplo.com.br",
             "", "autora", "R. das Flores, 22", "Belo Horizonte", "30130-000",
             "Exemplares de contrato", "12/05/2026", "Poeta, autora da casa"],
        ):
            sheet.append(row)
        for column, width in zip("ABCDEFGHIJK",
                                 (26, 24, 30, 20, 14, 28, 18, 12, 30, 14, 34)):
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"

        # As instruções vão numa aba SEPARADA, e não num rodapé da aba de
        # dados: o teste do ida-e-volta flagrou que uma linha de recado no fim
        # da planilha entra como contato chamado "(apague estas linhas...)".
        # O leitor só olha a aba de dados, então o recado não pode encostar
        # nela.
        guide = book.create_sheet("Como usar")
        for line in (
            "Como preencher",
            "",
            "1. Apague as linhas de exemplo da aba Contatos e ponha as suas.",
            "2. A ordem das colunas não importa — o importador reconhece os",
            "   títulos, e também variações comuns (Contato, Celular, Veículo).",
            "3. Só Nome OU E-mail é obrigatório. Com e-mail, o importador acha",
            "   quem já está na base e não duplica o cadastro.",
            "3b. A coluna Lista é opcional. Preenchida, um arquivo só alimenta",
            "   quantas listas você quiser — a lista que não existir é criada,",
            "   e a mesma pessoa pode repetir em quantas listas fizer sentido.",
            "   Vazia (ou ausente), tudo vai para a lista escolhida na tela.",
            "4. Endereço, Cidade e CEP não são obrigatórios aqui, mas sem os",
            "   três o pacote não sai do depósito na hora de enviar.",
            "5. Tipo aceita: autor, jornalista, influencer, livraria, outro.",
            "6. Último envio é opcional e é uma data DA LISTA, não do contato:",
            "   repita a mesma em todas as linhas do título (vale a maior).",
            "   Serve para ver na tela de Listas quais já esfriaram.",
            "7. Quem é: uma linha dizendo quem a pessoa é no mundo do livro",
            "   ('Crítica literária da Folha'). Vai para o cadastro do contato",
            "   e aparece ao lado do nome na hora de decidir um envio. Não",
            "   apaga o que já estiver escrito no contato.",
            "",
            "Ao subir a planilha, o importador mostra uma prévia do que vai",
            "acontecer antes de gravar qualquer coisa.",
        ):
            guide.append([line])
        guide["A1"].font = Font(bold=True, size=13)
        guide.column_dimensions["A"].width = 78

        stream = io.BytesIO()
        book.save(stream)
        return stream.getvalue()

    def action_back(self):
        self.ensure_one()
        self.write({'state': 'upload', 'preview_html': False})
        return self._reopen()

    def _reopen(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
