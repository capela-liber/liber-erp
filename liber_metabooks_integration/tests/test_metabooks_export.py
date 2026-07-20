# -*- coding: utf-8 -*-
"""Sending changes back to Metabooks.

Nothing here touches the network: the export produces a file and stops. What is
worth asserting is the loop a person actually depends on -- edit a book, see
exactly that change queued, get it out as a spreadsheet with the changed cell in
red, and have the book leave the queue only once the file was really delivered.
"""

import io
import zipfile
from datetime import date, timedelta

import openpyxl

from odoo import fields
from odoo.exceptions import UserError
from odoo.tests import TransactionCase, tagged

from ..services import metabooks_sheet as sheet


def _max_track(env):
    """Where a book's change history starts for a book already sent."""
    return env['mail.tracking.value'].search([], order='id desc', limit=1).id or 0


def _font_rgb(cell):
    """Colour of a cell as a string.

    openpyxl hands back a plain str for an explicit colour and an RGB
    descriptor object for a default one, so normalise before comparing.
    """
    colour = cell.font.color
    return str(colour.rgb) if colour is not None and colour.rgb else ''


@tagged('post_install', '-at_install')
class TestMetabooksExport(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.availability = cls.env['metabooks.avalaibility.definition'].create({
            'identify_number': '21',
            'product_definition': 'Disponível',
        })
        cls.role_author = cls.env['author.contributor.role'].search(
            [('name', '=', 'A01')], limit=1)
        if not cls.role_author:
            cls.role_author = cls.env['author.contributor.role'].create(
                {'name': 'A01'})
        cls.author = cls.env['metabooks.auther.publiser'].create({
            'name': 'Augusto da',
            'author_last_name': 'Silva',
            'author_contributor_role': cls.role_author.id,
        })
        cls.book = cls.env['product.template'].create({
            'name': 'O Cortiço',
            'barcode': '9788599296264',
            'list_price': 59.90,
            'metabooks_vendor_id': 'BR0089701',
            'metabooks_book_title': 'O Cortiço',
            'metabooks_page_count': 300,
            'metabooks_ncm': '4901.99.00',
            'metabooks_publish_date': date(2020, 5, 1),
            'metabooks_product_availability': cls.availability.id,
            'book_auther_ids': [(6, 0, cls.author.ids)],
        })
        # A book Metabooks already knows: that is what makes it a V (update)
        # rather than a Z (new title), and it is the cut-off for the history.
        cls.book.with_context(metabooks_from_sync=True).write({
            'metabooks_export_pending': False,
            'metabooks_export_pending_since': False,
            'metabooks_export_last': fields.Datetime.now() - timedelta(days=30),
            'metabooks_export_last_track': _max_track(cls.env),
        })
        cls._settle(cls.env)

    def _batch(self):
        return self.env['metabooks.export.batch'].create({})

    @classmethod
    def _settle(cls, env=None):
        """Close the books on the current transaction's tracking.

        Odoo writes tracking values from a precommit callback and keeps one set
        of "initial values" per record per transaction. In real use a book is
        created in one request and edited in another, so an edit is compared
        against the stored record. In a test everything shares a transaction:
        without settling in between, an edit is compared against the record as
        it was created, and the before/after comes out blank.
        """
        env = env or cls.env
        env.flush_all()
        env.cr.precommit.run()

    def _edit(self, records=None, **vals):
        """Edit a book the way the chatter will actually see it."""
        (records or self.book).write(vals)
        self._settle(self.env)

    # ------------------------------------------------------------------ #
    #  The queue
    # ------------------------------------------------------------------ #

    def test_editing_a_mapped_field_queues_the_book(self):
        self.book.metabooks_book_title = 'O Cortiço (edição comentada)'
        self.assertTrue(self.book.metabooks_export_pending)
        self.assertTrue(self.book.metabooks_export_pending_since)

    def test_editing_an_unmapped_field_leaves_the_book_alone(self):
        # description_sale is not a Metabooks column, so there is nothing to
        # send and queueing the book would be a false promise.
        self.book.description_sale = 'texto interno'
        self.assertFalse(self.book.metabooks_export_pending)

    def test_import_from_metabooks_does_not_queue_the_book(self):
        """The echo guard: their own data must not bounce back at them."""
        self.book.with_context(metabooks_from_sync=True).write(
            {'metabooks_book_title': 'Título vindo da Metabooks'})
        self.assertFalse(self.book.metabooks_export_pending)

    def test_clearing_pending_by_hand(self):
        self.book.metabooks_page_count = 310
        self.assertTrue(self.book.metabooks_export_pending)
        self.book.action_metabooks_clear_pending()
        self.assertFalse(self.book.metabooks_export_pending)

    # ------------------------------------------------------------------ #
    #  Gathering the diff
    # ------------------------------------------------------------------ #

    def test_prepare_records_before_and_after(self):
        self._edit(metabooks_page_count=320)

        batch = self._batch()
        batch.action_prepare()

        self.assertEqual(len(batch.line_ids), 1)
        line = batch.line_ids
        self.assertEqual(line.gtin, '9788599296264')
        self.assertEqual(batch.mb_id, 'BR0089701')

        pages = line.change_ids.filtered(
            lambda c: c.column == 'Número de páginas')
        self.assertEqual(len(pages), 1)
        self.assertEqual(pages.old_value, '300')
        self.assertEqual(pages.new_value, '320')

    def test_prepare_only_brings_what_changed(self):
        self._edit(metabooks_page_count=320)

        batch = self._batch()
        batch.action_prepare()

        columns = set(batch.change_ids.mapped('column'))
        self.assertEqual(columns, {'Número de páginas'},
                         "a book that only changed its page count should not "
                         "carry its whole record")

    def test_several_edits_collapse_to_first_and_last(self):
        """A -> B -> C is reported as A -> C: that is what they need to know."""
        self._edit(metabooks_book_title='Segundo título')
        self._edit(metabooks_book_title='Terceiro título')

        batch = self._batch()
        batch.action_prepare()
        change = batch.change_ids.filtered(lambda c: c.column == 'Título')
        self.assertEqual(change.old_value, 'O Cortiço')
        self.assertEqual(change.new_value, 'Terceiro título')

    def test_a_second_round_carries_only_the_new_change(self):
        """Where the history cut-off earns its keep.

        Reading the history from a timestamp lost this: mail.message.date only
        resolves to the second, so an edit landing in the same second as the
        send it followed looked older than that send, no history was found, and
        the fallback quietly sent every mapped column instead of the one that
        moved.
        """
        self._edit(metabooks_page_count=320)
        first = self._batch()
        first.action_prepare()
        first.action_generate()
        first.action_mark_sent()

        self._edit(metabooks_book_title='Título novo')
        second = self._batch()
        second.action_prepare()

        self.assertFalse(second.line_ids.no_history)
        self.assertEqual(set(second.change_ids.mapped('column')), {'Título'})
        change = second.change_ids
        self.assertEqual(change.old_value, 'O Cortiço')
        self.assertEqual(change.new_value, 'Título novo')

    def test_nothing_pending_is_an_error_not_an_empty_file(self):
        batch = self._batch()
        with self.assertRaises(UserError):
            batch.action_prepare()

    # ------------------------------------------------------------------ #
    #  Checking
    # ------------------------------------------------------------------ #

    def test_check_rejects_a_short_isbn(self):
        self._edit(metabooks_page_count=320)
        batch = self._batch()
        batch.action_prepare()
        batch.line_ids.gtin = '123'
        with self.assertRaises(UserError):
            batch.action_check()

    def test_check_rejects_two_books_with_the_same_isbn(self):
        twin = self.book.copy({'name': 'Cópia', 'barcode': '9788599296265'})
        twin.with_context(metabooks_from_sync=True).write({
            'metabooks_export_pending': False,
            'metabooks_export_last': fields.Datetime.now() - timedelta(days=30),
            'metabooks_export_last_track': _max_track(self.env),
        })
        self._settle(self.env)
        self._edit(self.book | twin, metabooks_page_count=321)

        batch = self._batch()
        batch.action_prepare()
        batch.line_ids[1].gtin = batch.line_ids[0].gtin
        with self.assertRaises(UserError):
            batch.action_check()

    def test_clearing_a_field_is_refused_not_silently_dropped(self):
        """A blank cell means "leave this alone" to Metabooks.

        So a field cleared in Odoo would reach them as no change at all, and the
        batch would report success over a change that never happened.
        """
        self._edit(metabooks_keywords='literatura, brasil')
        first = self._batch()
        first.action_prepare()
        first.action_generate()
        first.action_mark_sent()

        self._edit(metabooks_keywords=False)
        batch = self._batch()
        batch.action_prepare()
        with self.assertRaises(UserError):
            batch.action_check()
        with self.assertRaises(UserError):
            batch.action_generate()

    def test_generate_revalidates_an_already_checked_batch(self):
        self._edit(metabooks_page_count=320)
        batch = self._batch()
        batch.action_prepare()
        batch.action_check()
        batch.line_ids.gtin = '123'
        with self.assertRaises(UserError):
            batch.action_generate()

    def test_check_passes_and_moves_state(self):
        self._edit(metabooks_page_count=320)
        batch = self._batch()
        batch.action_prepare()
        batch.action_check()
        self.assertEqual(batch.state, 'checked')

    # ------------------------------------------------------------------ #
    #  The spreadsheet
    # ------------------------------------------------------------------ #

    def test_generated_sheet_carries_the_change_in_red(self):
        self._edit(metabooks_page_count=320)

        batch = self._batch()
        batch.action_prepare()
        batch.action_generate()

        self.assertEqual(batch.state, 'generated')
        self.assertTrue(batch.attachment_id)
        self.assertEqual(
            batch.filename,
            'V_BR0089701_%s_Alteracoes.xlsx' % date.today().strftime('%Y%m%d'))

        book = openpyxl.load_workbook(
            io.BytesIO(batch.attachment_id.raw))
        page = book.active
        headers = [c.value for c in page[1]]

        # Only what a human needs to read the row, plus what changed. Not 74
        # columns of blank.
        self.assertEqual(headers, ['GTIN', 'Título', 'Número de páginas'])

        row = {h: page.cell(row=2, column=i + 1)
               for i, h in enumerate(headers)}
        self.assertEqual(row['GTIN'].value, '9788599296264')
        self.assertEqual(row['Número de páginas'].value, 320)

        # The point of the exercise: changed cells stand out, the anchor does not.
        self.assertIn('B00020', _font_rgb(row['Número de páginas']))
        self.assertNotIn('B00020', _font_rgb(row['Título']))
        self.assertNotIn('B00020', _font_rgb(row['GTIN']))

    def test_dates_go_in_as_dates_not_text(self):
        """Their own template carries 44545, an Excel serial."""
        self._edit(metabooks_publish_date=date(2021, 3, 15))

        batch = self._batch()
        batch.action_prepare()
        batch.action_generate()

        page = openpyxl.load_workbook(
            io.BytesIO(batch.attachment_id.raw)).active
        headers = [c.value for c in page[1]]
        col = headers.index('Data de publicação') + 1
        self.assertEqual(page.cell(row=2, column=col).value.date(),
                         date(2021, 3, 15))

    def test_author_is_surname_first(self):
        self.assertEqual(self.book._metabooks_cell('Autor'),
                         'Silva, Augusto da')

    def test_availability_goes_as_its_label(self):
        self.assertEqual(self.book._metabooks_cell('Status de disponibilidade'),
                         'Disponível')

    # ------------------------------------------------------------------ #
    #  Delivery
    # ------------------------------------------------------------------ #

    def test_book_leaves_the_queue_only_once_marked_sent(self):
        self._edit(metabooks_page_count=320)

        batch = self._batch()
        batch.action_prepare()
        batch.action_generate()
        self.assertTrue(
            self.book.metabooks_export_pending,
            "generating a file is not delivering it -- an upload can fail")

        batch.action_mark_sent()
        self.assertEqual(batch.state, 'sent')
        self.assertFalse(self.book.metabooks_export_pending)
        self.assertTrue(self.book.metabooks_export_last)

    def test_unselected_books_stay_in_the_queue(self):
        twin = self.book.copy({'name': 'Cópia', 'barcode': '9788599296265'})
        twin.with_context(metabooks_from_sync=True).write({
            'metabooks_export_pending': False,
            'metabooks_export_last': fields.Datetime.now() - timedelta(days=30),
            'metabooks_export_last_track': _max_track(self.env),
        })
        self._settle(self.env)
        self._edit(self.book | twin, metabooks_page_count=321)

        batch = self._batch()
        batch.action_prepare()
        dropped = batch.line_ids.filtered(lambda l: l.product_id == twin)
        dropped.selected = False
        batch.action_generate()
        batch.action_mark_sent()

        self.assertFalse(self.book.metabooks_export_pending)
        self.assertTrue(twin.metabooks_export_pending,
                        "a book left unchecked was never sent, so it must stay")

    def test_a_sent_batch_is_frozen(self):
        self._edit(metabooks_page_count=320)
        batch = self._batch()
        batch.action_prepare()
        batch.action_generate()
        batch.action_mark_sent()
        with self.assertRaises(UserError):
            batch.action_reset()
        with self.assertRaises(UserError):
            batch.action_cancel()

    def test_history_survives_the_book_changing_again(self):
        self._edit(metabooks_page_count=320)
        batch = self._batch()
        batch.action_prepare()
        batch.action_generate()
        batch.action_mark_sent()

        self._edit(metabooks_page_count=999)

        change = batch.change_ids.filtered(
            lambda c: c.column == 'Número de páginas')
        self.assertEqual(change.new_value, '320',
                         "the log records what we sent, not what the book "
                         "happens to say later")


@tagged('post_install', '-at_install')
class TestMetabooksSheet(TransactionCase):
    """The file format on its own -- no products involved."""

    def test_filename_follows_the_2025_guide(self):
        name = sheet.build_filename(
            sheet.TASK_NEW, 'BR5108985', date(2025, 3, 20), 'Novos Títulos')
        self.assertEqual(name, 'Z_BR5108985_20250320_NovosTitulos.xlsx')

    def test_filename_rejects_an_unknown_task(self):
        with self.assertRaises(ValueError):
            sheet.build_filename('Q', 'BR1', date(2025, 1, 1), 'x')

    def test_unknown_column_is_refused(self):
        with self.assertRaises(ValueError):
            sheet.write_workbook(
                [{'values': {'Coluna Inventada': 'x'}, 'changed': set()}])

    def test_mandatory_columns_are_all_real_columns(self):
        self.assertEqual(set(sheet.MANDATORY) - set(sheet.COLUMNS), set())

    def test_workbook_is_a_valid_xlsx_with_one_sheet(self):
        data = sheet.write_workbook([{
            'values': {'GTIN': '9788599296264', 'Título': 'Um livro'},
            'changed': {'Título'},
        }])
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(data)))
        book = openpyxl.load_workbook(io.BytesIO(data))
        self.assertEqual(book.sheetnames, ['Planilha1'])

    def test_all_columns_emits_the_full_template(self):
        data = sheet.write_workbook(
            [{'values': {'GTIN': '1'}, 'changed': set()}], all_columns=True)
        page = openpyxl.load_workbook(io.BytesIO(data)).active
        self.assertEqual([c.value for c in page[1]], list(sheet.COLUMNS))
